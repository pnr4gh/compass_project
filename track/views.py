from django.shortcuts import render, redirect
from django.contrib.auth.decorators import user_passes_test, login_required
from . models import Assignment,ProblemStats, Solved_Problem, Assignment_Problems, Course, Course_Enrollement, ProblemStats, ProblemType, Contest, Platform, Problem, ProblemTags, User_Handle, Solved_Problem, Profile, Batch, Institution, Department, Complexity, Tags
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from openpyxl import load_workbook
from django.contrib.auth.models import User
from .forms import ProfileForm, UserCreationForm
from django.contrib.auth.hashers import make_password
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from .core.portal import get_leetcode_problems, get_g4g_problems, get_code_forces_problems, update_leet, update_code_forces, update_g4g
from django.contrib.auth import login, authenticate
from django.contrib.auth import logout
import leetcode as lc  
from django.db.models import Count, Q, F, ExpressionWrapper, DecimalField
from django.db import models
import requests
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework import status
from .serializers import ContestResultSerializer, ContestSerializer
from collections import defaultdict
from django.utils.timezone import localtime, now
from itertools import zip_longest

# API views
class ContestAPI(APIView):
    def post(self, request):
        serializer = ContestSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({"message": "Contest data received successfully!"}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
class ContestResultView(APIView):
    def post(self, request, *args, **kwargs):
        serializer = ContestResultSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({"message": "Contest data stored successfully"}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# Views
def contest_list(request):
    contests = Contest.objects.all()
    return render(request, "contest_list.html", {"contests": contests})

@login_required
def contest_stats(request):
    if request.user.is_superuser:
        users = User.objects.all()
        contests = Contest.objects.all()

        selected_user = request.GET.get("user")
        selected_contest = request.GET.get("contest")

        # Query the data based on filters
        query = ProblemStats.objects.all()
        if selected_user:
            query = query.filter(user_id=selected_user)
        if selected_contest:
            query = query.filter(contest_id=selected_contest)

        # Group data by contest
        grouped_stats = defaultdict(list)
        for stat in query:
            contest_name = stat.contest.name if stat.contest else "No Contest"
            grouped_stats[contest_name].append(stat)

        context = {
            'grouped_stats': dict(grouped_stats),
            'users': users,
            'contests': contests,
            'selected_user': int(selected_user) if selected_user else None,
            'selected_contest': int(selected_contest) if selected_contest else None,
        }

        return render(request, 'decodeschool/contest_stats.html', context)

    else:
        user_contest_stats = ProblemStats.objects.filter(user=request.user)
        # Grouping data by contest (handling None case)
        grouped_stats = defaultdict(list)
        for stat in user_contest_stats:
            contest_name = stat.contest.name if stat.contest else "No Contest"
            grouped_stats[contest_name].append(stat)

        return render(request, 'decodeschool/contest_stats.html', {'grouped_stats': dict(grouped_stats)})

def check_admin(user):
   return user.is_superuser

def check_coordinator(user):
    return user.groups.filter(name='coordinator').exists() or user.is_superuser

def logout_view(request):
    logout(request)
    return redirect('login')

def signup(request):
        
    if request.method == 'POST':
        user_form = UserCreationForm(request.POST)
        profile_form = ProfileForm(request.POST)

        if user_form.is_valid() and profile_form.is_valid():
            # Save the user (UserCreationForm)
            user = user_form.save()
            profile = profile_form.save(commit=False)
            profile.user = user
            profile.save()
           # Authenticate the user and log them in
            username = user_form.cleaned_data.get('username')
            raw_password = user_form.cleaned_data.get('password1')
            user = authenticate(username=username, password=raw_password)
            login(request, user)

            dmoj_register_url = f"{settings.DMOJ_API_BASE_URL}/accounts/api/register/"
            dmoj_data = {
                "username": username,
                "email": user.email,
                "password": raw_password,
                "organization": getattr(profile.institute, 'name', "Default Organization")

            }
            try:
                response = requests.post(dmoj_register_url, json=dmoj_data)
                # print("DMOJ API Response:", response.status_code, response.text)
                if response.status_code == 201:
                    messages.success(request, "Signup successful! You are also registered in DMOJ.")
                else:
                    messages.warning(request, "Signup successful, but DMOJ registration failed.")
            except requests.RequestException as e:
                # print(f"DMOJ registration error: {e}")  # Add this line
                messages.warning(request, f"DMOJ registration error: {e}")
            return redirect('dashboard')
    else:
        user_form = UserCreationForm()
        profile_form = ProfileForm()

    return render(request, 'decodeschool/signup.html', {
        'user_form': user_form,
        'profile_form' : profile_form
    })

@login_required
def change_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Important!
            messages.success(request, 'Your password was successfully updated!')
            return redirect('home')
        else:
            messages.error(request, 'Please correct the error below.')
    else:
        form = PasswordChangeForm(request.user)
    return render(request, 'decodeschool/change_password.html', {
        'form': form
    })

@login_required
def home(request):
    if check_coordinator(request.user):
        return redirect('myadmin')
    else:
        
        return redirect('dashboard')

@user_passes_test(check_coordinator)
def myadmin(request):
    user_tags = {}
    matching_problem_stats = []
    seen_user_problem_types = set()
    user_integrity_index = {}  # Will store average integrity_index per user

    solved_problems = Solved_Problem.objects.all()
    '''
    for solved_problem in solved_problems:
        user = solved_problem.user
        tags = solved_problem.get_tags()
        
        if user not in user_tags:
            user_tags[user] = set()
            user_integrity_index[user] = {'sum': 0, 'count': 0}  # Initialize sum and count
        
        user_tags[user].update(tags)

        problem_stats = ProblemStats.objects.filter(
            user=user,
            problem_type__name__in=tags,
        )

        for stat in problem_stats:
            problem_tags = ProblemTags.objects.filter(problem=solved_problem.problem)
            stat_tags = [tag.topictags.tag for tag in problem_tags]
            
            if set(tags).intersection(set(stat_tags)):
                user_problem_key = (user.id, stat.problem_type.name)
                
                if user_problem_key not in seen_user_problem_types:
                    seen_user_problem_types.add(user_problem_key)
                    matching_problem_stats.append(stat)
                    
                    # Sum the integrity_index values and increment count
                    user_integrity_index[user]['sum'] += getattr(stat, 'integrity_index', 0)
                    user_integrity_index[user]['count'] += 1

    # Calculate the average integrity_index for each user
    for user, data in user_integrity_index.items():
        average = data['sum'] / data['count'] if data['count'] > 0 else 0
        user_integrity_index[user] = average
        
        # Apply the logic to double the average if it is greater than 0.4
        if user_integrity_index[user] > 0.4:
            user_integrity_index[user] *= 2
    '''
    
    if request.headers.get('x-requested-with') == 'XMLHttpRequest': 
        department_id = request.GET.get('department')
        batch_id = request.GET.get('batch')
        institute_id = request.GET.get('institute')
        course_id = request.GET.get('course')

        queryset = Profile.objects.all()

        if department_id:
            queryset = queryset.filter(department__id=department_id)
        if batch_id:
            queryset = queryset.filter(batch__id=batch_id)
        if institute_id:
            queryset = queryset.filter(institute__id=institute_id)
        if course_id:
            queryset = queryset.filter(user__course_enrollement__course__id=course_id)

        leaderboard_data = queryset.annotate(total_solved=Count('user__solved_problem'),
        easy_solved=Count('user__solved_problem', filter=Q(user__solved_problem__problem__complexity__level="Easy")),
        medium_solved=Count('user__solved_problem', filter=Q(user__solved_problem__problem__complexity__level="Medium")),
        hard_solved=Count('user__solved_problem', filter=Q(user__solved_problem__problem__complexity__level="Hard"))
        ).annotate(total_score = ExpressionWrapper(F('total_solved') * F('integrity_index'), output_field=DecimalField())).order_by('-total_score')[0:20]

        data = list(leaderboard_data.values(
            'user__username', 'user__first_name', 'user__last_name',
            'department__short_name', 'batch__start_year', 'batch__end_year',
            'institute__name', 'easy_solved', 'medium_solved', 'hard_solved','total_score' ,'total_solved', 'integrity_index'
        ))
        '''
        for entry in data:
            user = User.objects.get(username=entry['user__username'])
            entry['user_integrity_index'] = user_integrity_index.get(user, 0)
        '''
        return JsonResponse(data, safe=False)

    platforms = Platform.objects.all()
    chart_data = []

    for platform in platforms:
        problems = Problem.objects.filter(platform=platform)
        complexity_counts = problems.values('complexity__level').annotate(count=models.Count('id'))
        
        chart_data.append({
            'object' : platform,
            'platform': platform.name,
            'data': list(complexity_counts)
        })

    context = {
        'departments': Department.objects.all(),
        'batches': Batch.objects.all(),
        'institutes': Institution.objects.all(),
        'courses': Course.objects.all(),
        'chart_data':chart_data,
        #'user_integrity_index': user_integrity_index,
       
    }
    return render(request, 'decodeschool/admin.html', context)

@login_required
def view_dashboard(request):
    user_handles = User_Handle.objects.filter(user=request.user)
    profile = Profile.objects.get(user=request.user)
    user = request.user
    today = timezone.now().date()
    
    # Filter assignments that have started and belong to courses the user is enrolled in
    assignments = Assignment.objects.filter(
        start_date__lte=today,
        course__course_enrollement__user=user
    )
    current_time = now()
    upcoming_contests = Contest.objects.filter(start_time__gt=current_time)
    live_contests = Contest.objects.filter(start_time__lte=current_time, end_time__gte=current_time)
    ended_contests = Contest.objects.filter(end_time__lt=current_time)

    for contest in upcoming_contests:
        contest.start_time = localtime(contest.start_time)
    for contest in live_contests:
        contest.start_time = localtime(contest.start_time)
        contest.end_time = localtime(contest.end_time)
    for contest in ended_contests:
        contest.end_time = localtime(contest.end_time)
    contests_zipped = list(zip_longest(upcoming_contests, live_contests, fillvalue=None))

    context = {
        'user_handles': user_handles,
        'profile': profile,
        'assignments': assignments,
        "upcoming_contests": upcoming_contests,
        "live_contests": live_contests,
        "ended_contests": ended_contests,
        "contests_zipped": contests_zipped,
    }
    return render(request, 'decodeschool/dashboard.html', context)


def contest_redirect(request,dmoj_key):
    contest_url = f"{settings.DMOJ_API_BASE_URL}/contest/{dmoj_key}/"
    return redirect(contest_url)

@user_passes_test(check_admin)
def get_portal_problems(request,portal_name):
    if portal_name == "Leet Code":
        get_leetcode_problems(request)
    elif portal_name == "Geeks for Geeks":
        get_g4g_problems(request)
    elif portal_name == "Code Forces":
        get_code_forces_problems(request)
    
    return redirect('myadmin')

@login_required
@csrf_exempt
def update_solved(request,uh,platform):
    data=dict()
    if platform == "Leet Code":
        
        return update_leet(request,uh)
    elif platform == "Geeks for Geeks":
        return update_g4g(request,uh)
  
    elif platform == "Code Forces":
        
        return update_code_forces(request,uh)
    return JsonResponse(data)

@user_passes_test(check_admin)
def import_from_excel(request):
    #return redirect('myadmin')
    data = dict()
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file)
        ws = wb.active
        data['count'] = 0
        data['missed'] = ""
        
        if(ws.max_column==2):
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:

                    username,course = row
                    if not Course_Enrollement.objects.filter(user=User.objects.get(username=username),course=Course.objects.get(course_title=course)).exists():

                        Course_Enrollement.objects.create(user=User.objects.get(username=username),course=Course.objects.get(course_title=course))
                    data['count'] += 1
                except:
                    data['missed'] += username + " "
        else:
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                
                username, first_name,last_name,email,mobile,department,batch,institution,leetcode = row
                try:
                    user = User.objects.create(username=username, first_name=first_name,last_name=last_name,password=make_password("Karpagam@105"),email=email)
                    Profile.objects.create(mobile_no=mobile,user=user,batch=Batch.objects.get(end_year=batch),department=Department.objects.get(short_name=department),institute=Institution.objects.get(name=institution))
                    User_Handle.objects.create(platform=Platform.objects.get(name="Leet Code"),user_handle=leetcode,user=user)
                    data['count'] += 1
                except Exception as e:
           
                    data['missed'] += username + " "
              
    return JsonResponse(data)


@login_required
def contest_and_solved_stats(request):
    user_tags = {}
    matching_problem_stats = []
    seen_user_problem_types = set()
    user_integrity_index = {}  # Will store average integrity_index per user

    solved_problems = Solved_Problem.objects.all()

    for solved_problem in solved_problems:
        user = solved_problem.user
        tags = solved_problem.get_tags()
        
        if user not in user_tags:
            user_tags[user] = set()
            user_integrity_index[user] = {'sum': 0, 'count': 0}  # Initialize sum and count
        
        user_tags[user].update(tags)

        problem_stats = ProblemStats.objects.filter(
            user=user,
            problem_type__name__in=tags,
        )

        for stat in problem_stats:
            problem_tags = ProblemTags.objects.filter(problem=solved_problem.problem)
            stat_tags = [tag.topictags.tag for tag in problem_tags]
            
            if set(tags).intersection(set(stat_tags)):
                user_problem_key = (user.id, stat.problem_type.name)
                
                if user_problem_key not in seen_user_problem_types:
                    seen_user_problem_types.add(user_problem_key)
                    matching_problem_stats.append(stat)
                    
                    # Sum the integrity_index values and increment count
                    user_integrity_index[user]['sum'] += getattr(stat, 'integrity_index', 0)
                    user_integrity_index[user]['count'] += 1

    # Calculate the average integrity_index for each user
    for user, data in user_integrity_index.items():
        average = data['sum'] / data['count'] if data['count'] > 0 else 0
        user_integrity_index[user] = average

        # Apply the logic to double the average if it is greater than 0.4
        if user_integrity_index[user] > 0.4:
            user_integrity_index[user] *= 2

    return render(request, 'decodeschool/contest_and_solved_stats.html', {
        'user_integrity_index': user_integrity_index,  # This now contains averages per user, with adjustments
    })

